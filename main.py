# Информационная система учета компьютерной техники

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import db_operations as db
from datetime import datetime, date
import os

# Настройка для поддержки русского языка в PDF
import sys
import io

# Устанавливаем кодировку для вывода в консоль
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# Попробуем импортировать библиотеки для отчетов
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # Регистрируем шрифты для поддержки русского языка
    font_paths = [
        'C:/Windows/Fonts/arial.ttf',
        'C:/Windows/Fonts/times.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        '/System/Library/Fonts/Arial.ttf',
    ]

    font_registered = False
    for font_path in font_paths:
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont('RussianFont', font_path))
                font_registered = True
                break
            except:
                continue

    if not font_registered:
        print("Внимание: Шрифт для русского языка не найден.")

    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class AuthWindow:
    """Окно авторизации"""

    def __init__(self):
        self.window = tk.Tk()
        self.window.title("Авторизация")
        self.window.geometry("300x220")
        self.window.resizable(False, False)
        self.window.eval('tk::PlaceWindow . center')

        self.login_var = tk.StringVar()
        self.password_var = tk.StringVar()
        self.create_widgets()

    def create_widgets(self):
        title = ttk.Label(self.window, text="Вход в систему", font=("Arial", 12, "bold"))
        title.pack(pady=15)

        frame = ttk.Frame(self.window, padding=15)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="Логин:").pack(anchor="w", pady=(0, 2))
        login_entry = ttk.Entry(frame, textvariable=self.login_var, width=30)
        login_entry.pack(fill="x", pady=(0, 10))
        login_entry.focus()

        ttk.Label(frame, text="Пароль:").pack(anchor="w", pady=(0, 2))
        password_entry = ttk.Entry(frame, textvariable=self.password_var, show="*", width=30)
        password_entry.pack(fill="x", pady=(0, 15))

        btn = ttk.Button(frame, text="Войти", command=self.login)
        btn.pack()

        self.window.bind('<Return>', lambda event: self.login())

    def login(self):
        login = self.login_var.get().strip()
        password = self.password_var.get().strip()

        if not login or not password:
            messagebox.showerror("Ошибка", "Введите логин и пароль")
            return

        user = db.authenticate_user(login, password)
        if user:
            user_data = {
                'user_id': user[0],
                'role': user[1],
                'employee_id': user[2],
                'full_name': f"{user[3]} {user[4]}"
            }
            self.window.destroy()
            MainApp(user_data)
        else:
            messagebox.showerror("Ошибка", "Неверный логин или пароль")
            self.login_var.set("")
            self.password_var.set("")

    def run(self):
        self.window.mainloop()


class MainApp:
    """Главное окно приложения"""

    def __init__(self, user_data):
        self.user_data = user_data
        self.is_admin = (user_data['role'] == 'admin')

        self.window = tk.Tk()
        self.window.title(f"Учет техники - {user_data['full_name']}")
        self.window.geometry("1100x700")

        self.create_menu()
        self.notebook = ttk.Notebook(self.window)
        self.notebook.pack(fill="both", expand=True, padx=5, pady=5)

        self.treeviews = {}
        self.search_configs = {}

        self.create_equipment_tab()
        self.create_employees_tab()
        self.create_movements_tab()
        self.create_repairs_tab()
        self.create_reports_tab()

        if self.is_admin:
            self.create_admin_tab()

        role_text = "Администратор" if self.is_admin else "Сотрудник техотдела"
        status = ttk.Label(self.window, text=f"Пользователь: {user_data['full_name']} | Роль: {role_text}",
                           relief="sunken", anchor="w")
        status.pack(side="bottom", fill="x")

        self.window.mainloop()

    def create_menu(self):
        menubar = tk.Menu(self.window)
        self.window.config(menu=menubar)

        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Сессия", menu=file_menu)
        file_menu.add_command(label="Сменить пользователя", command=self.logout)
        file_menu.add_separator()
        file_menu.add_command(label="Выход", command=self.window.destroy)

        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Справка", menu=help_menu)
        help_menu.add_command(label="О программе", command=self.show_about)

    def logout(self):
        self.window.destroy()
        AuthWindow().run()

    def show_about(self):
        messagebox.showinfo("О программе",
                            "Информационная система учета компьютерной техники\n\n"
                            "Возможности:\n"
                            "• Учет техники, сотрудников, перемещений и ремонтов\n"
                            "• Формирование отчетов с фильтрацией\n"
                            "• Сохранение отчетов в PDF и Excel\n"
                            "© 2026, ООО «ИТ Высота»")

    def apply_advanced_search(self, tab_name, data_func, field_names, all_data=None):
        config = self.search_configs.get(tab_name, {})
        search_field = config.get('field', '')
        search_condition = config.get('condition', 'contains')
        search_text = config.get('text', '').strip().lower()

        if not search_text or not search_field:
            self.refresh_table(tab_name, data_func)
            return

        if all_data is None:
            all_data = data_func()

        field_index = None
        for i, name in enumerate(field_names):
            if name == search_field:
                field_index = i
                break

        if field_index is None:
            self.refresh_table(tab_name, data_func)
            return

        filtered_data = []
        for row in all_data:
            if field_index >= len(row):
                continue
            cell_value = str(row[field_index]).lower() if row[field_index] else ""

            if search_condition == 'contains':
                if search_text in cell_value:
                    filtered_data.append(row)
            elif search_condition == 'starts_with':
                if cell_value.startswith(search_text):
                    filtered_data.append(row)
            elif search_condition == 'ends_with':
                if cell_value.endswith(search_text):
                    filtered_data.append(row)
            elif search_condition == 'equals':
                if cell_value == search_text:
                    filtered_data.append(row)
            elif search_condition == 'not_contains':
                if search_text not in cell_value:
                    filtered_data.append(row)

        tree = self.treeviews.get(tab_name)
        if not tree:
            return

        for item in tree.get_children():
            tree.delete(item)

        for row in filtered_data:
            tree.insert("", "end", values=row)

        count_label = self.treeviews.get(f"{tab_name}_count")
        if count_label:
            count_label.config(text=f"Записей: {len(filtered_data)}")

        return filtered_data

    def refresh_table(self, tab_name, data_func):
        tree = self.treeviews.get(tab_name)
        if not tree:
            return

        for item in tree.get_children():
            tree.delete(item)

        data = data_func()
        for row in data:
            tree.insert("", "end", values=row)

        count_label = self.treeviews.get(f"{tab_name}_count")
        if count_label:
            count_label.config(text=f"Записей: {len(data)}")

        return data

    def create_search_widgets(self, parent, tab_name, field_names, field_labels):
        search_frame = ttk.LabelFrame(parent, text="Расширенный поиск", padding=5)
        search_frame.pack(fill="x", padx=5, pady=5)

        ttk.Label(search_frame, text="Искать по:").grid(row=0, column=0, pady=3, padx=5, sticky="e")
        field_var = tk.StringVar()
        field_combo = ttk.Combobox(search_frame, textvariable=field_var, values=field_labels, width=20,
                                   state="readonly")
        field_combo.grid(row=0, column=1, pady=3, padx=5, sticky="w")
        if field_labels:
            field_combo.current(0)

        ttk.Label(search_frame, text="Условие:").grid(row=0, column=2, pady=3, padx=5, sticky="e")
        condition_var = tk.StringVar(value="contains")
        condition_combo = ttk.Combobox(search_frame, textvariable=condition_var,
                                       values=["contains", "starts_with", "ends_with", "equals", "not_contains"],
                                       width=12, state="readonly")
        condition_combo.grid(row=0, column=3, pady=3, padx=5, sticky="w")

        ttk.Label(search_frame, text="Значение:").grid(row=0, column=4, pady=3, padx=5, sticky="e")
        text_var = tk.StringVar()
        text_entry = ttk.Entry(search_frame, textvariable=text_var, width=25)
        text_entry.grid(row=0, column=5, pady=3, padx=5, sticky="w")

        def do_search():
            selected_label = field_var.get()
            for i, label in enumerate(field_labels):
                if label == selected_label:
                    field_name = field_names[i]
                    break
            else:
                field_name = field_names[0] if field_names else ""

            self.search_configs[tab_name] = {
                'field': field_name,
                'condition': condition_var.get(),
                'text': text_var.get()
            }
            self.apply_advanced_search(tab_name,
                                       lambda: self.get_all_data_for_tab(tab_name),
                                       field_names)

        def clear_search():
            field_var.set(field_labels[0] if field_labels else "")
            condition_var.set("contains")
            text_var.set("")
            self.search_configs[tab_name] = {}
            self.refresh_table(tab_name, lambda: self.get_all_data_for_tab(tab_name))

        ttk.Button(search_frame, text="Найти", command=do_search).grid(row=0, column=6, pady=3, padx=5)
        ttk.Button(search_frame, text="Сбросить", command=clear_search).grid(row=0, column=7, pady=3, padx=5)

    def get_all_data_for_tab(self, tab_name):
        if tab_name == 'equipment':
            return db.get_equipment()
        elif tab_name == 'employees':
            return db.get_employees()
        elif tab_name == 'movements':
            return db.get_movements(limit=1000)
        elif tab_name == 'repairs':
            return db.get_repairs()
        return []

    # ==================== ТЕХНИКА ====================

    def create_equipment_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Техника")

        toolbar = ttk.Frame(tab)
        toolbar.pack(fill="x", padx=5, pady=5)

        ttk.Button(toolbar, text="Добавить", command=self.add_equipment_dialog).pack(side="left", padx=2)
        ttk.Button(toolbar, text="Изменить статус", command=self.change_status_dialog).pack(side="left", padx=2)

        ttk.Button(toolbar, text="Обновить", command=lambda: self.refresh_table('equipment', db.get_equipment)).pack(
            side="left", padx=2)

        field_names = ["id_техники", "инвентарный_номер", "наименование", "модель", "серийный_номер", "статус"]
        field_labels = ["ID", "Инв.номер", "Наименование", "Модель", "Серийный номер", "Статус"]
        self.create_search_widgets(tab, 'equipment', field_names, field_labels)

        filter_frame = ttk.Frame(tab)
        filter_frame.pack(fill="x", padx=5, pady=5)

        ttk.Label(filter_frame, text="Статус:").pack(side="left", padx=5)
        statuses = [""] + [s[1] for s in db.get_statuses()]
        self.status_filter = ttk.Combobox(filter_frame, values=statuses, width=15)
        self.status_filter.pack(side="left", padx=5)

        ttk.Button(filter_frame, text="Применить фильтр", command=self.refresh_equipment_tree).pack(side="left", padx=5)

        columns = ("ID", "Инв.номер", "Наименование", "Модель", "Серийник", "Дата", "Статус")
        tree = ttk.Treeview(tab, columns=columns, show="headings", height=16)

        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=120)

        scroll = ttk.Scrollbar(tab, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scroll.set)

        tree.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        scroll.pack(side="right", fill="y", pady=5)

        count_label = ttk.Label(tab, text="Записей: 0", font=("Arial", 9))
        count_label.pack(side="bottom", pady=2)

        self.treeviews['equipment'] = tree
        self.treeviews['equipment_count'] = count_label
        self.refresh_equipment_tree()

    def refresh_equipment_tree(self):
        tree = self.treeviews.get('equipment')
        if not tree:
            return
        for item in tree.get_children():
            tree.delete(item)

        status = self.status_filter.get() if hasattr(self, 'status_filter') else None

        data = db.get_equipment(status if status else None)
        for row in data:
            tree.insert("", "end", values=row)

        count_label = self.treeviews.get('equipment_count')
        if count_label:
            count_label.config(text=f"Записей: {len(data)}")

        self.equipment_data = data

    def add_equipment_dialog(self):
        dialog = tk.Toplevel(self.window)
        dialog.title("Добавление техники")
        dialog.geometry("400x400")
        dialog.grab_set()
        dialog.resizable(False, False)

        ttk.Label(dialog, text="Инвентарный номер:").grid(row=0, column=0, pady=5, padx=10, sticky="e")
        inv_entry = ttk.Entry(dialog, width=30)
        inv_entry.grid(row=0, column=1, pady=5, padx=10)

        ttk.Label(dialog, text="Наименование:").grid(row=1, column=0, pady=5, padx=10, sticky="e")
        name_entry = ttk.Entry(dialog, width=30)
        name_entry.grid(row=1, column=1, pady=5, padx=10)

        ttk.Label(dialog, text="Модель:").grid(row=2, column=0, pady=5, padx=10, sticky="e")
        model_entry = ttk.Entry(dialog, width=30)
        model_entry.grid(row=2, column=1, pady=5, padx=10)

        ttk.Label(dialog, text="Серийный номер:").grid(row=3, column=0, pady=5, padx=10, sticky="e")
        serial_entry = ttk.Entry(dialog, width=30)
        serial_entry.grid(row=3, column=1, pady=5, padx=10)

        ttk.Label(dialog, text="Дата поступления:").grid(row=4, column=0, pady=5, padx=10, sticky="e")
        date_entry = ttk.Entry(dialog, width=30)
        date_entry.insert(0, db.get_current_date())
        date_entry.grid(row=4, column=1, pady=5, padx=10)

        ttk.Label(dialog, text="Статус:").grid(row=5, column=0, pady=5, padx=10, sticky="e")
        statuses = db.get_statuses()
        status_var = tk.StringVar()
        status_combo = ttk.Combobox(dialog, textvariable=status_var, values=[s[1] for s in statuses], width=27)
        status_combo.grid(row=5, column=1, pady=5, padx=10)
        if statuses:
            status_combo.current(0)

        def save():
            inv = inv_entry.get().strip()
            name = name_entry.get().strip()
            if not inv or not name:
                messagebox.showerror("Ошибка", "Заполните инвентарный номер и наименование")
                return
            status_id = next(s[0] for s in statuses if s[1] == status_var.get())
            result = db.add_equipment(inv, name, model_entry.get().strip(), serial_entry.get().strip(),
                                      date_entry.get().strip(), status_id)
            if result:
                messagebox.showinfo("Успех", "Техника добавлена")
                dialog.destroy()
                self.refresh_equipment_tree()
                self.search_configs['equipment'] = {}
            else:
                messagebox.showerror("Ошибка", "Не удалось добавить")

        ttk.Button(dialog, text="Сохранить", command=save).grid(row=6, column=0, columnspan=2, pady=15)

    def change_status_dialog(self):
        equipment = db.get_equipment()
        if not equipment:
            messagebox.showinfo("Информация", "Нет техники")
            return

        dialog = tk.Toplevel(self.window)
        dialog.title("Изменение статуса")
        dialog.geometry("350x150")
        dialog.grab_set()
        dialog.resizable(False, False)

        ttk.Label(dialog, text="Техника:").grid(row=0, column=0, pady=5, padx=10, sticky="e")
        eq_values = [f"{e[1]} - {e[2]}" for e in equipment]
        eq_var = tk.StringVar()
        eq_combo = ttk.Combobox(dialog, textvariable=eq_var, values=eq_values, width=30)
        eq_combo.grid(row=0, column=1, pady=5, padx=10)

        ttk.Label(dialog, text="Новый статус:").grid(row=1, column=0, pady=5, padx=10, sticky="e")
        statuses = db.get_statuses()
        status_var = tk.StringVar()
        status_combo = ttk.Combobox(dialog, textvariable=status_var, values=[s[1] for s in statuses], width=30)
        status_combo.grid(row=1, column=1, pady=5, padx=10)

        def save():
            if not eq_var.get() or not status_var.get():
                messagebox.showerror("Ошибка", "Заполните все поля")
                return
            idx = eq_values.index(eq_var.get())
            eq_id = equipment[idx][0]
            status_id = next(s[0] for s in statuses if s[1] == status_var.get())
            db.update_equipment_status(eq_id, status_id)
            messagebox.showinfo("Успех", "Статус изменен")
            dialog.destroy()
            self.refresh_equipment_tree()
            self.search_configs['equipment'] = {}

        ttk.Button(dialog, text="Сохранить", command=save).grid(row=2, column=0, columnspan=2, pady=10)

    # ==================== СОТРУДНИКИ ====================

    def create_employees_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Сотрудники")

        toolbar = ttk.Frame(tab)
        toolbar.pack(fill="x", padx=5, pady=5)

        if self.is_admin:
            ttk.Button(toolbar, text="Добавить", command=self.add_employee_dialog).pack(side="left", padx=2)

        ttk.Button(toolbar, text="Обновить", command=lambda: self.refresh_table('employees', db.get_employees)).pack(
            side="left", padx=2)

        field_names = ["id_сотрудника", "фамилия", "имя", "отчество", "должность", "отдел", "телефон"]
        field_labels = ["ID", "Фамилия", "Имя", "Отчество", "Должность", "Отдел", "Телефон"]
        self.create_search_widgets(tab, 'employees', field_names, field_labels)

        columns = ("ID", "Фамилия", "Имя", "Отчество", "Должность", "Отдел", "Телефон")
        tree = ttk.Treeview(tab, columns=columns, show="headings", height=16)

        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=120)

        scroll = ttk.Scrollbar(tab, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scroll.set)

        tree.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        scroll.pack(side="right", fill="y", pady=5)

        count_label = ttk.Label(tab, text="Записей: 0", font=("Arial", 9))
        count_label.pack(side="bottom", pady=2)

        self.treeviews['employees'] = tree
        self.treeviews['employees_count'] = count_label
        self.refresh_table('employees', db.get_employees)

    def add_employee_dialog(self):
        if not self.is_admin:
            return

        dialog = tk.Toplevel(self.window)
        dialog.title("Добавление сотрудника")
        dialog.geometry("350x350")
        dialog.grab_set()
        dialog.resizable(False, False)

        labels = ["Фамилия:", "Имя:", "Отчество:", "Должность:", "Отдел:", "Телефон:"]
        entries = []

        for i, label in enumerate(labels):
            ttk.Label(dialog, text=label).grid(row=i, column=0, pady=3, padx=10, sticky="e")
            entry = ttk.Entry(dialog, width=25)
            entry.grid(row=i, column=1, pady=3, padx=10)
            entries.append(entry)

        def save():
            values = [e.get().strip() for e in entries]
            if not values[0] or not values[1]:
                messagebox.showerror("Ошибка", "Фамилия и имя обязательны")
                return
            db.add_employee(values[0], values[1], values[2], values[3], values[4], values[5])
            messagebox.showinfo("Успех", "Сотрудник добавлен")
            dialog.destroy()
            self.refresh_table('employees', db.get_employees)
            self.search_configs['employees'] = {}

        ttk.Button(dialog, text="Сохранить", command=save).grid(row=len(labels), column=0, columnspan=2, pady=10)

    # ==================== ПЕРЕМЕЩЕНИЯ ====================

    def create_movements_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Перемещения")

        toolbar = ttk.Frame(tab)
        toolbar.pack(fill="x", padx=5, pady=5)

        ttk.Button(toolbar, text="Добавить", command=self.add_movement_dialog).pack(side="left", padx=2)
        ttk.Button(toolbar, text="Обновить",
                   command=lambda: self.refresh_table('movements', lambda: db.get_movements(limit=1000))).pack(
            side="left", padx=2)
        ttk.Button(toolbar, text="Сохранить форму", command=self.print_movement_form).pack(side="left", padx=2)

        field_names = ["id_перемещения", "инвентарный_номер", "наименование", "сотрудник", "дата"]
        field_labels = ["ID", "Инв.номер", "Наименование", "Сотрудник", "Дата"]
        self.create_search_widgets(tab, 'movements', field_names, field_labels)

        columns = ("ID", "Инв.номер", "Наименование", "Сотрудник", "Дата")
        tree = ttk.Treeview(tab, columns=columns, show="headings", height=16)

        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=150)

        scroll = ttk.Scrollbar(tab, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scroll.set)

        tree.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        scroll.pack(side="right", fill="y", pady=5)

        count_label = ttk.Label(tab, text="Записей: 0", font=("Arial", 9))
        count_label.pack(side="bottom", pady=2)

        self.treeviews['movements'] = tree
        self.treeviews['movements_count'] = count_label
        self.refresh_table('movements', lambda: db.get_movements(limit=1000))

    def add_movement_dialog(self):
        equipment = db.get_equipment()
        active = [e for e in equipment if e[6] != "списана"]
        if not active:
            messagebox.showinfo("Информация", "Нет доступной техники")
            return

        employees = db.get_employees()
        if not employees:
            messagebox.showinfo("Информация", "Нет сотрудников")
            return

        dialog = tk.Toplevel(self.window)
        dialog.title("Добавление перемещения")
        dialog.geometry("400x200")
        dialog.grab_set()
        dialog.resizable(False, False)

        ttk.Label(dialog, text="Техника:").grid(row=0, column=0, pady=5, padx=10, sticky="e")
        eq_values = [f"{e[1]} - {e[2]}" for e in active]
        eq_var = tk.StringVar()
        eq_combo = ttk.Combobox(dialog, textvariable=eq_var, values=eq_values, width=30)
        eq_combo.grid(row=0, column=1, pady=5, padx=10)

        ttk.Label(dialog, text="Сотрудник:").grid(row=1, column=0, pady=5, padx=10, sticky="e")
        emp_values = [f"{e[1]} {e[2]}" for e in employees]
        emp_var = tk.StringVar()
        emp_combo = ttk.Combobox(dialog, textvariable=emp_var, values=emp_values, width=30)
        emp_combo.grid(row=1, column=1, pady=5, padx=10)

        ttk.Label(dialog, text="Дата:").grid(row=2, column=0, pady=5, padx=10, sticky="e")
        date_entry = ttk.Entry(dialog, width=30)
        date_entry.insert(0, db.get_current_date())
        date_entry.grid(row=2, column=1, pady=5, padx=10)

        def save():
            if not eq_var.get() or not emp_var.get():
                messagebox.showerror("Ошибка", "Заполните все поля")
                return
            eq_idx = eq_values.index(eq_var.get())
            eq_id = active[eq_idx][0]
            emp_idx = emp_values.index(emp_var.get())
            emp_id = employees[emp_idx][0]

            db.add_movement(eq_id, emp_id, date_entry.get().strip())


            statuses = db.get_statuses()
            for s in statuses:
                if s[1] == "в работе":
                    db.update_equipment_status(eq_id, s[0])
                    break

            dialog.destroy()
            self.refresh_table('movements', lambda: db.get_movements(limit=1000))
            self.refresh_equipment_tree()

        ttk.Button(dialog, text="Сохранить", command=save).grid(row=3, column=0, columnspan=2, pady=10)

    def print_movement_form(self):
        """Печать акта передачи техники (форма для подписи сотрудника)"""
        if not PDF_AVAILABLE:
            messagebox.showerror("Ошибка", "Библиотека reportlab не установлена.\nУстановите: pip install reportlab")
            return

        tree = self.treeviews.get('movements')
        if not tree:
            messagebox.showwarning("Предупреждение", "Таблица перемещений не найдена")
            return
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("Предупреждение", "Выберите запись перемещения для печати")
            return
        item = tree.item(selected[0])
        mov_id = item['values'][0]
        mov = db.get_movement_by_id(mov_id)
        if not mov:
            messagebox.showerror("Ошибка", "Не удалось получить данные перемещения")
            return

        (mov_id, inv_num, tech_name, model, serial,
         surname, name, patronymic, position, department, date_str) = mov

        full_name = f"{surname} {name} {patronymic if patronymic else ''}".strip()

        filename = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=f"Акт_передачи_{inv_num}_{date_str}.pdf"
        )
        if not filename:
            return

        # Определяем шрифт с поддержкой русского языка (как в export_report_pdf)
        font_name = 'RussianFont' if font_registered else 'Helvetica'

        doc = SimpleDocTemplate(filename, pagesize=A4,
                                rightMargin=30, leftMargin=30,
                                topMargin=30, bottomMargin=30)
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=16,
            alignment=1,  # center
            spaceAfter=20,
            encoding='utf-8'
        )
        normal_style = ParagraphStyle(
            'NormalStyle',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=10,
            encoding='utf-8'
        )

        elements = []
        elements.append(Paragraph("Расписка о выдаче техники в личное пользование", title_style))
        elements.append(Spacer(1, 12))

        # Данные о передаче
        elements.append(Paragraph(f"<b>Дата передачи:</b> {date_str}", normal_style))
        elements.append(Spacer(1, 6))
        elements.append(Paragraph(f"<b>Инвентарный номер:</b> {inv_num}", normal_style))
        elements.append(Paragraph(f"<b>Наименование техники:</b> {tech_name}", normal_style))
        elements.append(Paragraph(f"<b>Модель:</b> {model if model else '—'}", normal_style))
        elements.append(Paragraph(f"<b>Серийный номер:</b> {serial if serial else '—'}", normal_style))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"<b>Сотрудник, получивший технику:</b> {full_name}", normal_style))
        elements.append(Paragraph(f"<b>Должность:</b> {position if position else '—'}", normal_style))
        elements.append(Paragraph(f"<b>Отдел:</b> {department if department else '—'}", normal_style))
        elements.append(Spacer(1, 30))

        # Таблица с подписями
        data = [
            ["", ""],
            ["Ответственный сотрудник технического отдела:", "Сотрудник, получивший технику:"],
            ["_________________________", "_________________________"],
            ["(подпись, расшифровка)", "(подпись, расшифровка)"]
        ]
        table = Table(data, colWidths=[230, 230])
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 1), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("Техника передана в исправном состоянии, претензий к работе не имею.", normal_style))

        doc.build(elements)
        messagebox.showinfo("Успех", f"Акт передачи сохранён:\n{filename}")
    # ==================== РЕМОНТЫ ====================

    def create_repairs_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Ремонты")

        toolbar = ttk.Frame(tab)
        toolbar.pack(fill="x", padx=5, pady=5)

        ttk.Button(toolbar, text="Добавить", command=self.add_repair_dialog).pack(side="left", padx=2)
        ttk.Button(toolbar, text="Обновить", command=lambda: self.refresh_table('repairs', db.get_repairs)).pack(
            side="left", padx=2)

        field_names = ["id_ремонта", "инвентарный_номер", "наименование", "дата_заявки", "дата_ремонта", "описание",
                       "стоимость", "статус"]
        field_labels = ["ID", "Инв.номер", "Наименование", "Дата заявки", "Дата ремонта", "Описание", "Стоимость",
                        "Статус"]
        self.create_search_widgets(tab, 'repairs', field_names, field_labels)

        columns = ("ID", "Инв.номер", "Наименование", "Дата заявки", "Дата ремонта", "Описание", "Стоимость", "Статус")
        tree = ttk.Treeview(tab, columns=columns, show="headings", height=13)

        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=100)
        tree.column("Описание", width=150)

        scroll = ttk.Scrollbar(tab, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scroll.set)

        tree.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        scroll.pack(side="right", fill="y", pady=5)

        count_label = ttk.Label(tab, text="Записей: 0", font=("Arial", 9))
        count_label.pack(side="bottom", pady=2)

        self.treeviews['repairs'] = tree
        self.treeviews['repairs_count'] = count_label
        self.refresh_table('repairs', db.get_repairs)

    def add_repair_dialog(self):
        equipment = db.get_equipment()
        if not equipment:
            messagebox.showinfo("Информация", "Нет техники")
            return

        dialog = tk.Toplevel(self.window)
        dialog.title("Добавление ремонта")
        dialog.geometry("450x400")
        dialog.grab_set()
        dialog.resizable(False, False)

        ttk.Label(dialog, text="Техника:").grid(row=0, column=0, pady=5, padx=10, sticky="e")
        eq_values = [f"{e[1]} - {e[2]}" for e in equipment]
        eq_var = tk.StringVar()
        eq_combo = ttk.Combobox(dialog, textvariable=eq_var, values=eq_values, width=30)
        eq_combo.grid(row=0, column=1, pady=5, padx=10)

        ttk.Label(dialog, text="Дата заявки:").grid(row=1, column=0, pady=5, padx=10, sticky="e")
        date_request = ttk.Entry(dialog, width=30)
        date_request.insert(0, db.get_current_date())
        date_request.grid(row=1, column=1, pady=5, padx=10)

        ttk.Label(dialog, text="Дата ремонта:").grid(row=2, column=0, pady=5, padx=10, sticky="e")
        date_repair = ttk.Entry(dialog, width=30)
        date_repair.grid(row=2, column=1, pady=5, padx=10)

        ttk.Label(dialog, text="Описание:").grid(row=3, column=0, pady=5, padx=10, sticky="e")
        desc_text = tk.Text(dialog, width=30, height=3)
        desc_text.grid(row=3, column=1, pady=5, padx=10)

        ttk.Label(dialog, text="Стоимость:").grid(row=4, column=0, pady=5, padx=10, sticky="e")
        cost_entry = ttk.Entry(dialog, width=30)
        cost_entry.grid(row=4, column=1, pady=5, padx=10)

        ttk.Label(dialog, text="Статус:").grid(row=5, column=0, pady=5, padx=10, sticky="e")
        status_combo = ttk.Combobox(dialog, values=["выполнен", "в работе", "отложен"], width=27)
        status_combo.grid(row=5, column=1, pady=5, padx=10)
        status_combo.current(1)

        def save():
            if not eq_var.get():
                messagebox.showerror("Ошибка", "Выберите технику")
                return
            idx = eq_values.index(eq_var.get())
            eq_id = equipment[idx][0]

            cost = cost_entry.get().strip()
            cost_val = float(cost) if cost else 0

            db.add_repair(eq_id, date_request.get().strip(),
                          date_repair.get().strip() if date_repair.get().strip() else None,
                          desc_text.get("1.0", "end-1c").strip(),
                          cost_val, status_combo.get())

            statuses = db.get_statuses()
            repair_status = next(s[0] for s in statuses if s[1] == "в ремонте")
            db.update_equipment_status(eq_id, repair_status)

            messagebox.showinfo("Успех", "Ремонт добавлен")
            dialog.destroy()
            self.refresh_table('repairs', db.get_repairs)
            self.refresh_equipment_tree()
            self.search_configs['repairs'] = {}

        ttk.Button(dialog, text="Сохранить", command=save).grid(row=6, column=0, columnspan=2, pady=10)

    # ==================== ОТЧЕТЫ ====================

    def create_reports_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Отчеты")

        self.report_type_names = {
            'equipment': 'Отчет по технике',
            'movements': 'Отчет по перемещениям',
            'repairs': 'Отчет по ремонтам',
            'assignment': 'Отчет по закреплению техники'
        }

        main_frame = ttk.Frame(tab, padding=10)
        main_frame.pack(fill="both", expand=True)

        # Тип отчета
        type_frame = ttk.LabelFrame(main_frame, text="Тип отчета", padding=10)
        type_frame.pack(fill="x", pady=(0, 10))

        self.report_type = tk.StringVar(value="equipment")

        ttk.Radiobutton(type_frame, text="Отчет по технике",
                        variable=self.report_type, value="equipment").pack(anchor="w", pady=2)
        ttk.Radiobutton(type_frame, text="Отчет по перемещениям",
                        variable=self.report_type, value="movements").pack(anchor="w", pady=2)
        ttk.Radiobutton(type_frame, text="Отчет по ремонтам",
                        variable=self.report_type, value="repairs").pack(anchor="w", pady=2)
        ttk.Radiobutton(type_frame, text="Отчет по закреплению техники",
                        variable=self.report_type, value="assignment").pack(anchor="w", pady=2)

        # Фильтры
        filter_frame = ttk.LabelFrame(main_frame, text="Фильтры (необязательно)", padding=10)
        filter_frame.pack(fill="x", pady=(0, 10))

        self.date_filter_frame = ttk.Frame(filter_frame)
        self.date_filter_frame.pack(fill="x", pady=5)

        ttk.Label(self.date_filter_frame, text="Период:").pack(side="left", padx=(0, 5))
        self.start_date = ttk.Entry(self.date_filter_frame, width=12)
        self.start_date.pack(side="left", padx=2)
        ttk.Label(self.date_filter_frame, text="—").pack(side="left", padx=2)
        self.end_date = ttk.Entry(self.date_filter_frame, width=12)
        self.end_date.pack(side="left", padx=2)
        ttk.Label(self.date_filter_frame, text="(ГГГГ-ММ-ДД)", foreground="gray").pack(side="left", padx=5)
        ttk.Button(self.date_filter_frame, text="Очистить", command=self.clear_dates, width=10).pack(side="left",
                                                                                                     padx=10)

        self.repair_status_filter_frame = ttk.Frame(filter_frame)
        ttk.Label(self.repair_status_filter_frame, text="Статус ремонта:").pack(side="left", padx=(0, 5))
        repair_statuses = ["", "в работе", "выполнен", "отложен"]
        self.repair_status = ttk.Combobox(self.repair_status_filter_frame, values=repair_statuses, width=20)
        self.repair_status.pack(side="left")

        self.status_filter_frame = ttk.Frame(filter_frame)
        ttk.Label(self.status_filter_frame, text="Статус техники:").pack(side="left", padx=(0, 5))
        statuses = [""] + [s[1] for s in db.get_statuses()]
        self.report_status = ttk.Combobox(self.status_filter_frame, values=statuses, width=20)
        self.report_status.pack(side="left")

        self.equipment_filter_frame = ttk.Frame(filter_frame)
        ttk.Label(self.equipment_filter_frame, text="Конкретная техника:").pack(side="left", padx=(0, 5))
        all_equipment = db.get_equipment()
        eq_values = [""] + [f"{e[1]} - {e[2]}" for e in all_equipment]
        self.report_equipment = ttk.Combobox(self.equipment_filter_frame, values=eq_values, width=30)
        self.report_equipment.pack(side="left")

        self.employee_filter_frame = ttk.Frame(filter_frame)
        ttk.Label(self.employee_filter_frame, text="Сотрудник:").pack(side="left", padx=(0, 5))
        all_employees = db.get_employees()
        emp_values = [""] + [f"{e[1]} {e[2]} {e[3]}" for e in all_employees]
        self.report_employee = ttk.Combobox(self.employee_filter_frame, values=emp_values, width=30)
        self.report_employee.pack(side="left")

        self.update_filters_visibility()

        # Кнопки
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=(0, 10))

        ttk.Button(button_frame, text="Сформировать отчет",
                   command=self.generate_report, width=20).pack(side="left", padx=5)
        ttk.Button(button_frame, text="Сохранить в Excel",
                   command=self.export_report_excel, width=20).pack(side="left", padx=5)
        ttk.Button(button_frame, text="Сохранить в PDF",
                   command=self.export_report_pdf, width=20).pack(side="left", padx=5)

        # Таблица результатов
        result_frame = ttk.LabelFrame(main_frame, text="Результат", padding=5)
        result_frame.pack(fill="both", expand=True)

        self.result_tree = ttk.Treeview(result_frame, show="headings")
        vsb = ttk.Scrollbar(result_frame, orient="vertical", command=self.result_tree.yview)
        hsb = ttk.Scrollbar(result_frame, orient="horizontal", command=self.result_tree.xview)
        self.result_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.result_tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        self.result_count = ttk.Label(result_frame, text="Записей: 0", font=("Arial", 9))
        self.result_count.pack(side="bottom", pady=5)

        self.report_type.trace('w', lambda *args: self.update_filters_visibility())

    def update_filters_visibility(self):
        rtype = self.report_type.get()

        if rtype in ['movements', 'repairs']:
            self.date_filter_frame.pack(fill="x", pady=5)
        else:
            self.date_filter_frame.pack_forget()

        if rtype == 'repairs':
            self.repair_status_filter_frame.pack(fill="x", pady=5)
        else:
            self.repair_status_filter_frame.pack_forget()

        if rtype == 'equipment':
            self.status_filter_frame.pack(fill="x", pady=5)
        else:
            self.status_filter_frame.pack_forget()

        if rtype != 'assignment':
            self.equipment_filter_frame.pack(fill="x", pady=5)
        else:
            self.equipment_filter_frame.pack_forget()

        if rtype == 'movements':
            self.employee_filter_frame.pack(fill="x", pady=5)
        else:
            self.employee_filter_frame.pack_forget()

    def clear_dates(self):
        self.start_date.delete(0, tk.END)
        self.end_date.delete(0, tk.END)

    def get_equipment_id_from_selection(self, equipment_str):
        if not equipment_str:
            return None
        parts = equipment_str.split(" - ")
        if len(parts) >= 2:
            inv_num = parts[0]
            all_eq = db.get_equipment()
            for eq in all_eq:
                if eq[1] == inv_num:
                    return eq[0]
        return None

    def get_employee_id_from_selection(self, employee_str):
        if not employee_str:
            return None
        all_emp = db.get_employees()
        for emp in all_emp:
            emp_full = f"{emp[1]} {emp[2]} {emp[3]}".strip()
            if emp_full == employee_str:
                return emp[0]
        return None

    def generate_report(self):
        """Формирование отчета"""
        rtype = self.report_type.get()

        # Получаем значения фильтров
        start = self.start_date.get().strip() if self.start_date.get().strip() else None
        end = self.end_date.get().strip() if self.end_date.get().strip() else None

        # Для отладки - выводим введенные даты
        print(f"Фильтр по датам: начало={start}, конец={end}")

        status = self.report_status.get() if hasattr(self, 'report_status') and self.report_status.get() else None
        repair_status = self.repair_status.get() if hasattr(self,
                                                            'repair_status') and self.repair_status.get() else None
        equipment_str = self.report_equipment.get() if hasattr(self,
                                                               'report_equipment') and self.report_equipment.get() else None
        employee_str = self.report_employee.get() if hasattr(self,
                                                             'report_employee') and self.report_employee.get() else None

        equipment_id = self.get_equipment_id_from_selection(equipment_str)
        employee_id = self.get_employee_id_from_selection(employee_str)

        if rtype == 'equipment':
            headers, data = self.get_equipment_report(status, equipment_id)
        elif rtype == 'movements':
            headers, data = self.get_movements_report(start, end, equipment_id, employee_id)
        elif rtype == 'repairs':
            headers, data = self.get_repairs_report(start, end, equipment_id, repair_status)
        elif rtype == 'assignment':
            headers, data = self.get_assignment_report()
        else:
            return

        # Очищаем предыдущие результаты
        self.result_tree.delete(*self.result_tree.get_children())

        if not data:
            self.result_tree["columns"] = []
            self.result_count.config(text="Записей: 0")
            messagebox.showinfo("Информация", "Нет данных за выбранный период")
            return

        # Настраиваем колонки
        self.result_tree["columns"] = list(range(len(headers)))
        for i, header in enumerate(headers):
            self.result_tree.heading(i, text=header)
            self.result_tree.column(i, width=120, minwidth=80)

        # Заполняем данными
        for row in data:
            display_row = []
            for val in row:
                if val is None:
                    display_row.append("")
                elif isinstance(val, (date, datetime)):
                    display_row.append(val.strftime('%Y-%m-%d'))
                else:
                    display_row.append(str(val))
            self.result_tree.insert("", "end", values=display_row)

        self.result_count.config(text=f"Записей: {len(data)}")

        # Сохраняем для экспорта
        self.current_report_headers = headers
        self.current_report_data = data
        self.current_report_type = self.report_type_names.get(rtype, rtype)

    def get_equipment_report(self, status, equipment_id):
        data = db.get_equipment(status)
        if equipment_id:
            data = [d for d in data if d[0] == equipment_id]
        headers = ["ID", "Инв.номер", "Наименование", "Модель", "Серийный номер", "Дата", "Статус"]
        return headers, data

    def get_movements_report(self, start, end, equipment_id, employee_id):
        """Отчет по перемещениям с корректной фильтрацией по датам"""
        movements = db.get_movements(limit=10000)

        print(f"movements_report: фильтрация по датам {start} - {end}")

        filtered = []
        for mov in movements:
            # mov: (id, инв_номер, наименование, сотрудник, дата)
            mov_date_value = mov[4] if len(mov) > 4 else None

            # Проверка фильтра по датам
            if start or end:
                if not mov_date_value:
                    continue  # Пропускаем записи без даты

                try:
                    # Преобразуем дату
                    if isinstance(mov_date_value, str):
                        mov_date = datetime.strptime(mov_date_value, '%Y-%m-%d').date()
                    elif isinstance(mov_date_value, date):
                        mov_date = mov_date_value
                    else:
                        continue

                    # Проверяем начальную дату
                    if start:
                        start_date = datetime.strptime(start, '%Y-%m-%d').date()
                        if mov_date < start_date:
                            continue

                    # Проверяем конечную дату
                    if end:
                        end_date = datetime.strptime(end, '%Y-%m-%d').date()
                        if mov_date > end_date:
                            continue

                except Exception as e:
                    print(f"Ошибка парсинга даты: {e}")
                    continue

            # Фильтр по технике
            if equipment_id:
                if mov[1] != self.get_inventory_number_by_id(equipment_id):
                    continue

            # Фильтр по сотруднику
            if employee_id:
                employee = db.get_employee_by_id(employee_id)
                if employee and f"{employee[1]} {employee[2]}" != mov[3]:
                    continue

            filtered.append(mov)

        print(f"Найдено перемещений: {len(filtered)} из {len(movements)}")
        headers = ["ID", "Инв.номер", "Наименование", "Сотрудник", "Дата"]
        return headers, filtered

    def get_repairs_report(self, start, end, equipment_id, repair_status):
        """Отчет по ремонтам"""
        # Вызываем функцию БД с передачей параметров дат
        # Убедитесь, что в db_operations.py функция get_repairs принимает period_start и period_end
        repairs = db.get_repairs(period_start=start, period_end=end)

        print(f"repairs_report: фильтрация по датам {start} - {end}, получено ремонтов: {len(repairs)}")

        filtered = []
        for repair in repairs:
            # repair: (id, инв_номер, наименование, дата_заявки, дата_ремонта, описание, стоимость, статус)

            # Фильтр по статусу ремонта
            if repair_status and len(repair) > 7 and repair[7] != repair_status:
                continue

            # Фильтр по технике
            if equipment_id:
                if repair[1] != self.get_inventory_number_by_id(equipment_id):
                    continue

            filtered.append(repair)

        print(f"После дополнительных фильтров: {len(filtered)}")
        headers = ["ID", "Инв.номер", "Наименование", "Дата заявки", "Дата ремонта", "Описание", "Стоимость", "Статус"]
        return headers, filtered

    def get_inventory_number_by_id(self, equipment_id):
        all_equipment = db.get_equipment()
        for eq in all_equipment:
            if eq[0] == equipment_id:
                return eq[1]
        return None

    def get_assignment_report(self):
        data = db.get_report_equipment_by_employee()
        headers = ["Сотрудник", "Должность", "Количество техники"]
        return headers, data

    def export_report_excel(self):
        if not EXCEL_AVAILABLE:
            messagebox.showerror("Ошибка", "Библиотека openpyxl не установлена.\nУстановите: pip install openpyxl")
            return

        if not hasattr(self, 'current_report_data') or not self.current_report_data:
            messagebox.showwarning("Предупреждение", "Сначала сформируйте отчет")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"{self.current_report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        if not filename:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Отчет"

            title = f"{self.current_report_type} от {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            ws.merge_cells(f'A1:{openpyxl.utils.get_column_letter(len(self.current_report_headers))}1')
            ws['A1'] = title
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal="center")

            for col, header in enumerate(self.current_report_headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            for row_idx, row in enumerate(self.current_report_data, 4):
                for col_idx, value in enumerate(row, 1):
                    if value is None:
                        ws.cell(row=row_idx, column=col_idx, value="")
                    elif isinstance(value, (date, datetime)):
                        ws.cell(row=row_idx, column=col_idx, value=value.strftime('%Y-%m-%d'))
                    else:
                        ws.cell(row=row_idx, column=col_idx, value=value)

            for col in range(1, len(self.current_report_headers) + 1):
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(col)
                for row in range(3, len(self.current_report_data) + 4):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                max_length = min(max(max_length, len(self.current_report_headers[col - 1])), 50)
                ws.column_dimensions[column_letter].width = max_length + 2

            wb.save(filename)
            messagebox.showinfo("Успех", f"Отчет сохранен:\n{filename}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить Excel:\n{str(e)}")

    def export_report_pdf(self):
        if not PDF_AVAILABLE:
            messagebox.showerror("Ошибка", "Библиотека reportlab не установлена.\nУстановите: pip install reportlab")
            return

        if not hasattr(self, 'current_report_data') or not self.current_report_data:
            messagebox.showwarning("Предупреждение", "Сначала сформируйте отчет")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=f"{self.current_report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        )

        if not filename:
            return

        try:
            from reportlab.lib.pagesizes import landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

            font_registered = False
            font_paths = [
                'C:/Windows/Fonts/arial.ttf',
                'C:/Windows/Fonts/times.ttf',
                '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                '/System/Library/Fonts/Arial.ttf',
            ]

            for font_path in font_paths:
                if os.path.exists(font_path):
                    try:
                        pdfmetrics.registerFont(TTFont('RussianFont', font_path))
                        font_registered = True
                        break
                    except:
                        continue

            font_name = 'RussianFont' if font_registered else 'Helvetica'

            doc = SimpleDocTemplate(filename, pagesize=landscape(A4),
                                    rightMargin=15, leftMargin=15,
                                    topMargin=20, bottomMargin=20)
            elements = []

            styles = getSampleStyleSheet()

            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Normal'],
                fontName=font_name,
                fontSize=14,
                alignment=1,
                spaceAfter=20,
                encoding='utf-8'
            )

            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontName=font_name,
                fontSize=10,
                encoding='utf-8'
            )

            title = f"{self.current_report_type} от {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            elements.append(Paragraph(title, title_style))
            elements.append(Spacer(1, 10))

            table_data = [self.current_report_headers]
            for row in self.current_report_data:
                row_data = []
                for val in row:
                    if val is None:
                        row_data.append("")
                    elif isinstance(val, (date, datetime)):
                        row_data.append(val.strftime('%Y-%m-%d'))
                    else:
                        row_data.append(str(val))
                table_data.append(row_data)

            col_count = len(self.current_report_headers)
            available_width = landscape(A4)[0] - 30
            col_width = available_width / col_count

            table = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)

            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.green),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))

            for i in range(1, len(table_data)):
                table.setStyle(TableStyle([('FONTNAME', (0, i), (-1, i), font_name)]))

            elements.append(table)
            elements.append(Spacer(1, 20))

            count_text = f"Всего записей: {len(self.current_report_data)}"
            elements.append(Paragraph(count_text, normal_style))

            doc.build(elements)
            messagebox.showinfo("Успех", f"Отчет сохранен:\n{filename}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить PDF:\n{str(e)}")

    # ==================== АДМИНИСТРИРОВАНИЕ ====================

    def create_admin_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Администрирование")

        frame = ttk.LabelFrame(tab, text="Пользователи", padding=10)
        frame.pack(fill="x", padx=10, pady=10)

        ttk.Button(frame, text="Список пользователей", command=self.show_users_list).pack(side="left", padx=5)
        ttk.Button(frame, text="Добавить пользователя", command=self.add_user_dialog).pack(side="left", padx=5)

        info_frame = ttk.LabelFrame(tab, text="Статистика", padding=10)
        info_frame.pack(fill="x", padx=10, pady=10)

        stats = f"Техники: {db.get_equipment_count()}\nСотрудников: {db.get_employees_count()}\nРемонтов: {db.get_repairs_count()}"
        ttk.Label(info_frame, text=stats, justify="left").pack(anchor="w")

    def show_users_list(self):
        users = db.get_all_users()
        if not users:
            messagebox.showinfo("Информация", "Нет зарегистрированных пользователей")
            return

        dialog = tk.Toplevel(self.window)
        dialog.title("Управление пользователями")
        dialog.geometry("700x450")
        dialog.grab_set()

        # Таблица со списком пользователей
        columns = ("ID", "Логин", "Роль", "Сотрудник")
        tree = ttk.Treeview(dialog, columns=columns, show="headings")
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=120)
        for user in users:
            tree.insert("", "end", values=user)

        tree.pack(fill="both", expand=True, padx=10, pady=10)

        # Фрейм для кнопок
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)

        # Получить ID выбранного пользователя
        def get_selected():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Предупреждение", "Выберите пользователя")
                return None
            item = tree.item(selected[0])
            return item['values'][0]

        # Редактирование
        def edit_user():
            user_id = get_selected()
            if user_id:
                self.edit_user_dialog(user_id, dialog)

        # Удаление
        def delete_user():
            user_id = get_selected()
            if user_id and messagebox.askyesno("Подтверждение", "Удалить выбранного пользователя?"):
                if db.delete_user(user_id):
                    messagebox.showinfo("Успех", "Пользователь удалён")
                    dialog.destroy()
                    self.show_users_list()  # обновить список
                else:
                    messagebox.showerror("Ошибка", "Не удалось удалить пользователя")

        ttk.Button(btn_frame, text="Изменить", command=edit_user).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Удалить", command=delete_user).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Закрыть", command=dialog.destroy).pack(side="left", padx=5)

        # Двойной клик для быстрого редактирования
        tree.bind("<Double-1>", lambda e: edit_user())

    def edit_user_dialog(self, user_id, parent_dialog):
        """Диалог редактирования пользователя"""
        user_data = db.get_user_by_id(user_id)
        if not user_data:
            messagebox.showerror("Ошибка", "Не удалось загрузить данные пользователя")
            return

        (uid, login, role, emp_id,
         surname, name, patronymic, position, department) = user_data

        employees = db.get_employees()
        if not employees:
            messagebox.showerror("Ошибка", "Нет сотрудников. Сначала добавьте сотрудника.")
            return

        dialog = tk.Toplevel(parent_dialog)
        dialog.title(f"Редактирование пользователя: {login}")
        dialog.geometry("500x400")
        dialog.grab_set()
        dialog.resizable(False, False)

        # --- Поля ввода ---
        ttk.Label(dialog, text="Логин:").grid(row=0, column=0, pady=5, padx=10, sticky="e")
        login_entry = ttk.Entry(dialog, width=25)
        login_entry.insert(0, login)
        login_entry.grid(row=0, column=1, pady=5, padx=10)

        ttk.Label(dialog, text="Новый пароль (оставьте пустым, чтобы не менять):").grid(
            row=1, column=0, pady=5, padx=10, sticky="e")
        pass_entry = ttk.Entry(dialog, width=25, show="*")
        pass_entry.grid(row=1, column=1, pady=5, padx=10)

        ttk.Label(dialog, text="Подтверждение пароля:").grid(row=2, column=0, pady=5, padx=10, sticky="e")
        confirm_entry = ttk.Entry(dialog, width=25, show="*")
        confirm_entry.grid(row=2, column=1, pady=5, padx=10)

        ttk.Label(dialog, text="Роль:").grid(row=3, column=0, pady=5, padx=10, sticky="e")
        role_combo = ttk.Combobox(dialog, values=["admin", "technician"], width=22)
        role_combo.set(role)
        role_combo.grid(row=3, column=1, pady=5, padx=10)

        ttk.Label(dialog, text="Сотрудник:").grid(row=4, column=0, pady=5, padx=10, sticky="e")
        emp_values = [f"{e[1]} {e[2]} ({e[4]})" for e in employees]
        emp_combo = ttk.Combobox(dialog, values=emp_values, width=22)
        # Найти индекс текущего сотрудника
        emp_index = -1
        for i, e in enumerate(employees):
            if e[0] == emp_id:
                emp_index = i
                break
        if emp_index >= 0:
            emp_combo.set(emp_values[emp_index])
        emp_combo.grid(row=4, column=1, pady=5, padx=10)

        # Рамка с требованиями к паролю
        requirements_frame = ttk.LabelFrame(dialog, text="Требования к паролю", padding=5)
        requirements_frame.grid(row=5, column=0, columnspan=2, pady=10, padx=10, sticky="ew")
        requirements_text = """
        • Минимальная длина: 12 символов
        • Заглавные и строчные буквы
        • Цифры
        • Специальные символы (!@#$%^&*)
        • Не должен быть слишком простым
        """
        ttk.Label(requirements_frame, text=requirements_text, justify="left", font=("Arial", 8)).pack(anchor="w")

        def save():
            new_login = login_entry.get().strip()
            new_password = pass_entry.get().strip()
            confirm = confirm_entry.get().strip()
            new_role = role_combo.get()
            new_employee_str = emp_combo.get()

            if not new_login:
                messagebox.showerror("Ошибка", "Логин не может быть пустым")
                return
            if new_password and new_password != confirm:
                messagebox.showerror("Ошибка", "Пароли не совпадают")
                return
            if new_password:
                is_valid, msg = db.validate_password_strength(new_password)
                if not is_valid:
                    messagebox.showerror("Ошибка", msg)
                    return
            if not new_role:
                messagebox.showerror("Ошибка", "Выберите роль")
                return
            if not new_employee_str:
                messagebox.showerror("Ошибка", "Выберите сотрудника")
                return

            try:
                idx = emp_values.index(new_employee_str)
                new_emp_id = employees[idx][0]
            except ValueError:
                messagebox.showerror("Ошибка", "Выберите сотрудника из списка")
                return

            success, msg = db.update_user(user_id, new_login, new_password, new_role, new_emp_id)
            if success:
                messagebox.showinfo("Успех", msg)
                dialog.destroy()
                parent_dialog.destroy()  # закрыть окно списка
                self.show_users_list()  # открыть обновлённый список
            else:
                messagebox.showerror("Ошибка", msg)

        ttk.Button(dialog, text="Сохранить", command=save).grid(row=6, column=0, columnspan=2, pady=15)

    def add_user_dialog(self):
        employees = db.get_employees()
        if not employees:
            messagebox.showerror("Ошибка", "Нет сотрудников. Сначала добавьте сотрудника.")
            return

        dialog = tk.Toplevel(self.window)
        dialog.title("Добавление пользователя")
        dialog.geometry("400x450")
        dialog.grab_set()
        dialog.resizable(False, False)

        dialog.transient(self.window)
        dialog.update_idletasks()
        x = self.window.winfo_x() + (self.window.winfo_width() - 400) // 2
        y = self.window.winfo_y() + (self.window.winfo_height() - 450) // 2
        dialog.geometry(f"+{x}+{y}")

        ttk.Label(dialog, text="Логин:").grid(row=0, column=0, pady=5, padx=10, sticky="e")
        login_entry = ttk.Entry(dialog, width=25)
        login_entry.grid(row=0, column=1, pady=5, padx=10)

        ttk.Label(dialog, text="Пароль:").grid(row=1, column=0, pady=5, padx=10, sticky="e")
        pass_entry = ttk.Entry(dialog, width=25, show="*")
        pass_entry.grid(row=1, column=1, pady=5, padx=10)

        ttk.Label(dialog, text="Подтверждение пароля:").grid(row=2, column=0, pady=5, padx=10, sticky="e")
        confirm_entry = ttk.Entry(dialog, width=25, show="*")
        confirm_entry.grid(row=2, column=1, pady=5, padx=10)

        ttk.Label(dialog, text="Роль:").grid(row=3, column=0, pady=5, padx=10, sticky="e")
        role_combo = ttk.Combobox(dialog, values=["admin", "technician"], width=22)
        role_combo.grid(row=3, column=1, pady=5, padx=10)
        role_combo.current(1)

        ttk.Label(dialog, text="Сотрудник:").grid(row=4, column=0, pady=5, padx=10, sticky="e")
        emp_values = [f"{e[1]} {e[2]} ({e[4]})" for e in employees]
        emp_combo = ttk.Combobox(dialog, values=emp_values, width=22)
        emp_combo.grid(row=4, column=1, pady=5, padx=10)

        requirements_frame = ttk.LabelFrame(dialog, text="Требования к паролю", padding=5)
        requirements_frame.grid(row=5, column=0, columnspan=2, pady=10, padx=10, sticky="ew")

        requirements_text = """
        • Минимальная длина: 12 символов
        • Заглавные и строчные буквы
        • Цифры
        • Специальные символы (!@#$%^&*)
        • Не должен быть слишком простым
        """
        ttk.Label(requirements_frame, text=requirements_text, justify="left", font=("Arial", 8)).pack(anchor="w")

        def save():
            login = login_entry.get().strip()
            password = pass_entry.get().strip()
            confirm = confirm_entry.get().strip()
            role = role_combo.get()
            employee = emp_combo.get()

            if not login:
                messagebox.showerror("Ошибка", "Введите логин")
                return
            if not password:
                messagebox.showerror("Ошибка", "Введите пароль")
                return
            if password != confirm:
                messagebox.showerror("Ошибка", "Пароли не совпадают")
                return
            if not role:
                messagebox.showerror("Ошибка", "Выберите роль")
                return
            if not employee:
                messagebox.showerror("Ошибка", "Выберите сотрудника")
                return

            try:
                idx = emp_values.index(employee)
                employee_id = employees[idx][0]
            except ValueError:
                messagebox.showerror("Ошибка", "Выберите сотрудника из списка")
                return

            success, message = db.add_user(login, password, role, employee_id)

            if success:
                messagebox.showinfo("Успех", message)
                dialog.destroy()
            else:
                messagebox.showerror("Ошибка", message)

        ttk.Button(dialog, text="Сохранить", command=save).grid(row=6, column=0, columnspan=2, pady=15)


if __name__ == "__main__":
    app = AuthWindow()
    app.run()